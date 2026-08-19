"""Import/export helpers: Excel, CSV, JSON, and PDF generation."""
import csv
import io
import json
from datetime import date, datetime
from decimal import Decimal


def _serialize(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def model_to_dict(obj, fields):
    return {f: _serialize(getattr(obj, f, None)) for f in fields}


def export_csv(rows, fields):
    """rows: list of dicts. Returns CSV bytes."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fields)
    writer.writeheader()
    for r in rows:
        writer.writerow(r)
    return buf.getvalue().encode("utf-8")


def export_json(rows):
    return json.dumps(rows, indent=2, default=_serialize).encode("utf-8")


def export_excel(rows, fields, sheet_name="Data"):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(fields)
    for r in rows:
        ws.append([r.get(f) for f in fields])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_rows(file_storage):
    """Read an uploaded CSV or Excel file into a list of dicts."""
    filename = (file_storage.filename or "").lower()
    if filename.endswith(".csv") or filename.endswith(".txt"):
        text = file_storage.read().decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    if filename.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        wb = load_workbook(file_storage, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(h) if h is not None else "" for h in rows[0]]
        return [dict(zip(header, r)) for r in rows[1:]]
    raise ValueError("Unsupported file type. Use CSV, TXT, or XLSX.")
